import pandas as pd
from datetime import datetime
from tkinter import Tk, filedialog, messagebox, Button, Label
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.datetime import from_excel
import os

def process_excel(file_path):
    
    wb = load_workbook(file_path, data_only=False)  # do NOT strip formulas or objects


    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)

    today = datetime.today().date()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue

                parsed_date = None
                if isinstance(value, datetime):
                    parsed_date = value.date()
                elif isinstance(value, (int, float)):
                    try:
                        parsed_date = from_excel(value).date()
                    except:
                        continue
                elif isinstance(value, str):
                    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
                    if not pd.isna(parsed):
                        parsed_date = parsed.date()

                if parsed_date is None:
                    continue

                days_left = (parsed_date - today).days
                if days_left < 0:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = red_font
                elif days_left <= 7:
                    cell.fill = red_fill
                    cell.font = Font(color="000000")
                elif days_left <= 14:
                    cell.fill = yellow_fill
                    cell.font = Font(color="000000")

    base, ext = os.path.splitext(file_path)
    output_file = base + "_Formatted" + ext
    wb.save(output_file)
    return output_file

def choose_file():
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        return
    try:
        output_file = process_excel(file_path)
        messagebox.showinfo("Success", f"File processed successfully!\nSaved as:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

def main():
    root = Tk()
    root.title("Deadline Highlighter")
    root.geometry("300x150")

    Label(root, text="Excel Deadline Highlighter", font=("Arial", 14, "bold")).pack(pady=10)
    Button(root, text="Select Excel File", command=choose_file, width=20).pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()
