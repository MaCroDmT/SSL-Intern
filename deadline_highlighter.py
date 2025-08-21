import pandas as pd
from datetime import datetime
from tkinter import Tk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.datetime import from_excel
import os

def process_excel(file_path):
    # Load workbook
    
    wb = load_workbook(file_path, data_only=True)  # do NOT strip formulas or objects

    # Define styles
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # yellow
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")    # light red
    red_font = Font(color="9C0006", bold=True)

    today = datetime.today().date()

    # Loop all sheets and cells
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue

                parsed_date = None

                # Case 1: Already datetime
                if isinstance(value, datetime):
                    parsed_date = value.date()

                # Case 2: Excel serial number
                elif isinstance(value, (int, float)):
                    try:
                        parsed_date = from_excel(value).date()
                    except:
                        continue

                # Case 3: String date like "6-Aug"
                elif isinstance(value, str):
                    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
                    if not pd.isna(parsed):
                        parsed_date = parsed.date()

                if parsed_date is None:
                    continue

                days_left = (parsed_date - today).days

                if days_left < 0:  # deadline passed → text red
                    cell.fill = PatternFill(fill_type=None)  # remove background
                    cell.font = red_font
                elif days_left <= 7:  # within 1 week → red cell
                    cell.fill = red_fill
                    cell.font = Font(color="000000")  # normal black text
                elif days_left <= 14:  # within 2 weeks → yellow cell
                    cell.fill = yellow_fill
                    cell.font = Font(color="000000")  # normal black text
                # else: untouched

    # Save output
    base, ext = os.path.splitext(file_path)
    output_file = base + "_Formatted" + ext
    wb.save(output_file)
    return output_file


def main():
    # Create file dialog
    root = Tk()
    root.withdraw()  # hide main window
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )

    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        return

    try:
        output_file = process_excel(file_path)
        messagebox.showinfo("Success", f"File processed successfully!\nSaved as:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")


if __name__ == "__main__":
    main()
